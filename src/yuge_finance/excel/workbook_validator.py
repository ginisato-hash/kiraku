"""Excel出力の検証（喜らく単体）。

- 出力ファイルが存在し、テンプレ本体とは別物であること。
- 保護シートの代表的な数式セルが数式のまま残っていること（値で潰されていない）。
- シート名が維持されていること。
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import openpyxl

from .. import config

# 保護シートで数式が残っているべき代表セル
_FORMULA_SENTINELS = {
    "05_試算表": ["D6", "E6"],
    "06_PL": [],
    "11_モデル連携": ["C6"],
    "13_チェック": ["C5", "E5"],
}
EXPECTED_SHEETS = [
    "00_ダッシュボード", "01_入力ページ", "02_銀行API取込", "03_OTA取込", "04_仕訳帳",
    "05_試算表", "06_PL", "07_BS", "08_CF", "09_借入返済", "10_KPI", "11_モデル連携",
    "12_勘定科目マスタ", "13_チェック", "14_使い方",
]


def validate(workbook_path: Path) -> List[Dict]:
    out: List[Dict] = []
    wb_path = Path(workbook_path)

    def add(name, ok, detail=""):
        out.append({"check": name, "status": "OK" if ok else "要確認", "detail": detail})

    add("出力ファイル存在", wb_path.exists(), str(wb_path))
    add("テンプレ未上書き", wb_path.resolve() != config.template_path().resolve())
    if not wb_path.exists():
        return out

    wb = openpyxl.load_workbook(str(wb_path), data_only=False)
    add("シート名維持", list(wb.sheetnames) == EXPECTED_SHEETS,
        f"{wb.sheetnames}")

    for sheet, cells in _FORMULA_SENTINELS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for addr in cells:
            v = ws[addr].value
            is_formula = isinstance(v, str) and v.startswith("=")
            add(f"数式保持 {sheet}!{addr}", is_formula, f"value={v!r}")
    return out
