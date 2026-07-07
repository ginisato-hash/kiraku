"""Excel出力（喜らく単体）。

テンプレを読み込み、書込可能シートのみにデータを書く。
- 06_PL/07_BS/08_CF/10_KPI/11_モデル連携 の数式セルは触らない。
- テンプレ本体は絶対に上書きしない（出力は data/output/<month>/）。
- A列(仕訳ID)・C列(月) は各行に数式を再設定し、試算表SUMIFSを維持。
- 書式・色・シート名は openpyxl がスタイルを保持。
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.workbook.properties import CalcProperties

from .. import config
from ..normalize.schema import BookingRecord, JournalEntry


def _to_date(s):
    if isinstance(s, (date, datetime)):
        return s
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except ValueError:
        return s


def _col_idx(letter: str) -> int:
    return openpyxl.utils.column_index_from_string(letter)


class WorkbookWriter:
    def __init__(self) -> None:
        self.map = config.load_yaml("workbook_map.yml")
        self.writable = set(self.map["writable_sheets"])
        self.protected = set(self.map["protected_sheets"])

    def _check_writable(self, sheet: str) -> None:
        if sheet in self.protected:
            raise PermissionError(f"保護シートへの書込みは禁止: {sheet}")
        if sheet not in self.writable:
            raise PermissionError(f"書込み許可されていないシート: {sheet}")

    def _write_table(self, wb, key: str, records: List[dict]) -> int:
        spec = self.map["sheets"][key]
        sheet = spec["sheet"]
        self._check_writable(sheet)
        ws = wb[sheet]
        start = spec["start_row"]
        cols = spec["columns"]
        formula_cols = spec.get("formula_columns", {})
        n = 0
        for i, rec in enumerate(records):
            row = start + i
            for field, letter in cols.items():
                val = rec.get(field, "")
                if field.endswith("_date") or field == "transaction_date":
                    val = _to_date(val)
                ws.cell(row=row, column=_col_idx(letter)).value = val
            for letter, tmpl in formula_cols.items():
                ws.cell(row=row, column=_col_idx(letter)).value = tmpl.format(row=row)
            n += 1
        return n

    def write(self, month: str,
              journal: List[JournalEntry],
              bank_rows: List[dict],
              beds24_rows: List[dict],
              loan_rollforward: Dict = None,
              checks: List[Dict] = None,
              revenue_recon: Dict = None,
              output_path: Path = None) -> Path:
        tpl = config.template_path()
        if not tpl.exists():
            raise FileNotFoundError(f"テンプレが見つかりません: {tpl}")
        wb = openpyxl.load_workbook(str(tpl), data_only=False)

        # 銀行明細
        self._write_table(wb, "銀行", bank_rows)
        # Beds24予約（OTA取込）
        self._write_table(wb, "beds24", beds24_rows)
        # 仕訳帳
        jrows = [self._journal_to_row(e) for e in journal]
        self._write_table(wb, "仕訳帳", jrows)

        # 試算表 対象月
        ym = month.replace("-", "/")
        tb_spec = self.map["sheets"]["試算表"]
        self._check_writable(tb_spec["sheet"])
        wb[tb_spec["sheet"]][tb_spec["cells"]["target_month"]] = ym

        # 借入返済 実績ブロック
        if loan_rollforward:
            self._write_loan_block(wb, month, loan_rollforward)
        # チェック 実績ブロック
        if checks:
            self._write_check_block(wb, checks)
        # 売上 reconciliation ブロック
        if revenue_recon:
            self._write_revenue_recon_block(wb, revenue_recon)

        # 開いたとき Excel/LibreOffice が再計算するようにする
        if wb.calculation is None:
            wb.calculation = CalcProperties(fullCalcOnLoad=True)
        else:
            wb.calculation.fullCalcOnLoad = True

        out = output_path or (config.output_dir(month) / "updated_workbook.xlsx")
        if Path(out).resolve() == tpl.resolve():
            raise PermissionError("テンプレート本体への上書きは禁止です。")
        wb.save(str(out))
        return Path(out)

    @staticmethod
    def _journal_to_row(e: JournalEntry) -> dict:
        return {
            "journal_date": e.journal_date,
            "description": e.description,
            "debit_account": e.debit_account, "debit_subaccount": e.debit_subaccount,
            "debit_amount": e.debit_amount,
            "credit_account": e.credit_account, "credit_subaccount": e.credit_subaccount,
            "credit_amount": e.credit_amount,
            "tax_category": e.tax_category, "counterparty": e.counterparty,
            "source": f"{e.source}/{e.rule_id}",
            "memo": e.memo,
        }

    def _write_loan_block(self, wb, month: str, lr: Dict) -> None:
        spec = self.map["loan_block"]
        self._check_writable(spec["sheet"])
        ws = wb[spec["sheet"]]
        r = spec["start_row"]
        ws.cell(r, 1).value = spec["header"][0]
        ws.cell(r, 2).value = month
        ws.cell(r + 1, 1).value = "期首残高"
        ws.cell(r + 1, 2).value = lr["opening"]
        ws.cell(r + 2, 1).value = "新規借入"
        ws.cell(r + 2, 2).value = lr["新規借入"]
        ws.cell(r + 3, 1).value = "元本返済"
        ws.cell(r + 3, 2).value = lr["元本返済"]
        ws.cell(r + 4, 1).value = "期末残高"
        ws.cell(r + 4, 2).value = lr["closing"]

    def _write_check_block(self, wb, checks: List[Dict]) -> None:
        spec = self.map["check_block"]
        self._check_writable(spec["sheet"])
        ws = wb[spec["sheet"]]
        r = spec["start_row"]
        for j, h in enumerate(spec["header"]):
            ws.cell(r, 1 + j).value = h
        for i, c in enumerate(checks, start=1):
            ws.cell(r + i, 1).value = c["check"]
            ws.cell(r + i, 2).value = str(c["value"])
            ws.cell(r + i, 3).value = str(c["allow"])
            ws.cell(r + i, 4).value = c["status"]

    def _write_revenue_recon_block(self, wb, rec: Dict) -> None:
        from ..accounting import revenue_recon as rr_mod
        spec = self.map.get("revenue_recon_block")
        if not spec:
            return
        self._check_writable(spec["sheet"])
        ws = wb[spec["sheet"]]
        r = spec["start_row"]
        ws.cell(r, 1).value = spec.get("title", "売上reconciliation")
        for j, h in enumerate(spec["header"]):
            ws.cell(r + 1, 1 + j).value = h
        for i, ln in enumerate(rr_mod.display_lines(rec), start=2):
            ws.cell(r + i, 1).value = ln["区分"]
            ws.cell(r + i, 2).value = ln["項目"]
            ws.cell(r + i, 3).value = ln["値"]
